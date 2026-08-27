import io
from datetime import datetime
import openpyxl
import streamlit as st

st.set_page_config(
    page_title="Order Data Synchronizer",
    page_icon="📊",
    layout="wide"
)

# Key Columns and Row Offsets
DATA_KEY_COL = "K"          # OrdNo in DATA file
DATA_START_ROW = 4          # Data begins on Row 4

WORKDATA_KEY_COL = "A"      # OrdNo in WorkDataNew file
WORKDATA_START_ROW = 4      # Data begins on Row 4

# Target (DATA) -> Source (WorkDataNew) Coordinates
COLUMN_MAPPING = {
    "N": "CG",  # Total Bag Bal
    "O": "DJ",  # Bal To Exp Qty
    "P": "DQ",  # PSEGOMA
    "Q": "DX",  # GSIR
    "R": "DY",  # GSIT
    "S": "EA",  # CastBal
    "T": "EP",  # RHODIUM
    "U": "EZ",  # P3POLA
    "V": "FA",  # P3POLB
    "W": "FB",  # P3POLC
    "X": "FE",  # P4POLB
    "Y": "FF",  # P1SETA
    "Z": "FJ",  # SETP
    "AA": "FO", # P1SETFK
    "AB": "GA", # P3PPLA
    "AC": "GF", # P4PPLB
    "AD": "GJ", # PFMG
    "AE": "GM", # P1FILC
    "AF": "GX", # P4FILB
    "AG": "GY", # SAMPLE
    "AH": "HA", # OS
    "AI": "HU", # TACHE
    "AJ": "JI", # SPRU
    "AK": "KV"  # BalToOpnQty
}


def clean_order_id(value):
    """Clean and normalize Order ID string."""
    if value is None:
        return ""
    val_str = str(value).strip()
    if val_str.endswith(".0"):
        val_str = val_str[:-2]
    return val_str


st.title("📊 Daily Order Data Synchronizer")
st.markdown("Upload both daily Excel files to synchronize your maintained **DATA** sheet.")

st.divider()

# Two-column layout for file uploads
col1, col2 = st.columns(2)

with col1:
    st.subheader("1. Maintained Template")
    file_data = st.file_uploader("Upload DATA.xlsx (Target file to update)", type=["xlsx"])

with col2:
    st.subheader("2. System Export")
    file_workdata = st.file_uploader("Upload WorkDataNew.xlsx (Source data)", type=["xlsx"])

st.divider()

# Process action
if file_data and file_workdata:
    if st.button("🚀 Run Update Process", type="primary", use_container_width=True):
        with st.spinner("Processing in-memory... Please wait."):
            try:
                # 1. Load source workbook (read calculated values)
                file_workdata.seek(0)
                wb_source = openpyxl.load_workbook(file_workdata, data_only=True)
                ws_source = wb_source.active

                # 2. Index WorkDataNew records
                source_lookup = {}
                for row_idx in range(WORKDATA_START_ROW, ws_source.max_row + 1):
                    raw_id = ws_source[f"{WORKDATA_KEY_COL}{row_idx}"].value
                    ord_id = clean_order_id(raw_id)
                    if not ord_id:
                        continue
                    if ord_id not in source_lookup:
                        source_lookup[ord_id] = {
                            src_col: ws_source[f"{src_col}{row_idx}"].value
                            for src_col in COLUMN_MAPPING.values()
                        }

                # 3. Load target DATA workbook (preserve formulas and cell styling)
                file_data.seek(0)
                wb_target = openpyxl.load_workbook(file_data, data_only=False)
                ws_target = wb_target.active

                updated_count = 0
                missing_orders = []

                # 4. In-place update starting from Row 4
                for row_idx in range(DATA_START_ROW, ws_target.max_row + 1):
                    raw_id = ws_target[f"{DATA_KEY_COL}{row_idx}"].value
                    ord_id = clean_order_id(raw_id)
                    if not ord_id:
                        continue

                    if ord_id in source_lookup:
                        for target_col, src_col in COLUMN_MAPPING.items():
                            ws_target[f"{target_col}{row_idx}"].value = source_lookup[ord_id].get(src_col)
                        updated_count += 1
                    else:
                        missing_orders.append({"Row in DATA": row_idx, "Order ID": ord_id})

                # 5. Save updated workbook directly to memory buffer
                output_buffer = io.BytesIO()
                wb_target.save(output_buffer)
                output_buffer.seek(0)

                # 6. Display metrics and download action
                st.success("✨ Update completed successfully!")
                
                m1, m2 = st.columns(2)
                m1.metric("Orders Updated", updated_count)
                m2.metric("Unmatched / Flagged Orders", len(missing_orders))

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                download_name = f"DATA_Updated_{timestamp}.xlsx"

                st.download_button(
                    label="📥 Download Updated DATA.xlsx",
                    data=output_buffer,
                    file_name=download_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

                if missing_orders:
                    st.warning(f"⚠️ {len(missing_orders)} orders in DATA were not found in WorkDataNew:")
                    st.dataframe(missing_orders, use_container_width=True)

            except Exception as e:
                st.error(f"Processing error: {str(e)}")
else:
    st.info("💡 Please upload both Excel files above to proceed.")