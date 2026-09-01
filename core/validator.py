"""
Pre-flight validator for Excel workbooks and schema integrity.
"""

import io
from typing import Dict, Any, List
import openpyxl
from core.exceptions import (
    CorruptedFileError,
    InsufficientDataError,
    MissingColumnsError,
    SwappedFilesError
)
from core.utils import clean_order_id, col_to_index


def safe_load_workbook(file_bytes: io.BytesIO, data_only: bool = False, file_label: str = "Uploaded file") -> openpyxl.Workbook:
    """Safely attempts to load an Excel workbook from an in-memory byte buffer."""
    try:
        file_bytes.seek(0)
        wb = openpyxl.load_workbook(file_bytes, data_only=data_only)
        return wb
    except Exception as e:
        raise CorruptedFileError(
            f"{file_label} is invalid, corrupted, or password-protected. "
            f"Please ensure it is an unencrypted .xlsx file. (Details: {str(e)})"
        )


def validate_sync_inputs(
    ws_data: openpyxl.worksheet.worksheet.Worksheet,
    ws_workdata: openpyxl.worksheet.worksheet.Worksheet,
    profile: Dict[str, Any]
) -> List[str]:
    """
    Performs pre-flight structural checks on both worksheets.
    
    Returns:
        List[str]: Non-blocking advisory warnings detected during validation.
    """
    warnings: List[str] = []

    data_start_row = profile["data_start_row"]
    workdata_start_row = profile["workdata_start_row"]
    data_key_col = profile["data_key_col"]
    workdata_key_col = profile["workdata_key_col"]
    mapping = profile["column_mapping"]

    # 1. Row count validation
    if ws_data.max_row < data_start_row:
        raise InsufficientDataError(
            f"The Maintained DATA file has only {ws_data.max_row} row(s). "
            f"Expected data to start at row {data_start_row}."
        )

    if ws_workdata.max_row < workdata_start_row:
        raise InsufficientDataError(
            f"The System Export (WorkDataNew) file has only {ws_workdata.max_row} row(s). "
            f"Expected data to start at row {workdata_start_row}."
        )

    # 2. Maximum column boundary verification
    max_src_col_idx = max(col_to_index(c) for c in mapping.values())
    if ws_workdata.max_column < max_src_col_idx:
        missing_cols = [
            src for src in mapping.values()
            if col_to_index(src) > ws_workdata.max_column
        ]
        raise MissingColumnsError(
            f"The System Export file is missing required columns. "
            f"The sheet ends at column {openpyxl.utils.get_column_letter(ws_workdata.max_column)}, "
            f"but mapping requires: {', '.join(missing_cols[:5])}..."
        )

    # 3. Swapped File Heuristic Detection
    # Sample Order IDs from both files
    data_samples = [
        clean_order_id(ws_data[f"{data_key_col}{r}"].value)
        for r in range(data_start_row, min(data_start_row + 5, ws_data.max_row + 1))
        if ws_data[f"{data_key_col}{r}"].value is not None
    ]
    
    work_col_a_samples = [
        clean_order_id(ws_workdata[f"{workdata_key_col}{r}"].value)
        for r in range(workdata_start_row, min(workdata_start_row + 5, ws_workdata.max_row + 1))
        if ws_workdata[f"{workdata_key_col}{r}"].value is not None
    ]

    work_col_k_samples = [
        clean_order_id(ws_workdata[f"{data_key_col}{r}"].value)
        for r in range(workdata_start_row, min(workdata_start_row + 5, ws_workdata.max_row + 1))
        if ws_workdata[f"{data_key_col}{r}"].value is not None
    ]

    # If DATA file has empty keys in Col K, but has keys in Col A, files may be swapped
    if not any(data_samples) and any(work_col_a_samples):
        raise SwappedFilesError(
            "Column K in the Maintained DATA file has no order numbers. "
            "Please check if you accidentally uploaded the files in the wrong upload boxes."
        )

    # If WorkDataNew has no keys in Col A, but has keys in Col K
    if not any(work_col_a_samples) and any(work_col_k_samples):
        raise SwappedFilesError(
            "Column A in the System Export file has no order numbers. "
            "It appears the System Export and Maintained DATA files were swapped."
        )

    return warnings