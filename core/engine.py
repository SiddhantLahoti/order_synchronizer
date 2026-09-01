"""
Core synchronization engine with validation, duplicate detection, and execution audit.
"""

import io
from typing import Dict, Any, Tuple, List
import openpyxl
from core.validator import safe_load_workbook, validate_sync_inputs
from core.exceptions import NoMatchingOrdersError
from core.utils import clean_order_id


def process_sync(
    data_file_bytes: io.BytesIO,
    workdata_file_bytes: io.BytesIO,
    profile: Dict[str, Any]
) -> Tuple[io.BytesIO, Dict[str, Any]]:
    """
    Executes pre-flight validation and in-memory cell synchronization.

    Returns:
        output_buffer (io.BytesIO): In-memory updated Excel file buffer.
        audit_report (Dict[str, Any]): Execution summary containing counts, missing orders,
                                       duplicate entries, and warnings.
    """
    # 1. Safe Load
    wb_source = safe_load_workbook(workdata_file_bytes, data_only=True, file_label="System Export (WorkDataNew)")
    wb_target = safe_load_workbook(data_file_bytes, data_only=False, file_label="Maintained DATA template")

    ws_source = wb_source.active
    ws_target = wb_target.active

    # 2. Run Pre-flight Validation
    advisory_warnings = validate_sync_inputs(ws_target, ws_source, profile)

    src_key_col = profile["workdata_key_col"]
    src_start_row = profile["workdata_start_row"]
    target_key_col = profile["data_key_col"]
    target_start_row = profile["data_start_row"]
    column_mapping = profile["column_mapping"]

    # 3. Build Source Lookup Index & Track Duplicates
    source_lookup: Dict[str, Dict[str, Any]] = {}
    duplicate_source_orders: List[str] = []

    for row_idx in range(src_start_row, ws_source.max_row + 1):
        raw_id = ws_source[f"{src_key_col}{row_idx}"].value
        ord_id = clean_order_id(raw_id)
        
        if not ord_id:
            continue

        if ord_id in source_lookup:
            if ord_id not in duplicate_source_orders:
                duplicate_source_orders.append(ord_id)
        else:
            source_lookup[ord_id] = {
                src_col: ws_source[f"{src_col}{row_idx}"].value
                for src_col in column_mapping.values()
            }

    if duplicate_source_orders:
        advisory_warnings.append(
            f"Detected {len(duplicate_source_orders)} duplicate order(s) in the system export. "
            f"The first occurrence was used for mapping."
        )

    # 4. Perform In-Place Updates
    updated_count = 0
    missing_orders: List[Dict[str, Any]] = []

    for row_idx in range(target_start_row, ws_target.max_row + 1):
        raw_id = ws_target[f"{target_key_col}{row_idx}"].value
        ord_id = clean_order_id(raw_id)
        
        if not ord_id:
            continue

        if ord_id in source_lookup:
            for target_col, src_col in column_mapping.items():
                ws_target[f"{target_col}{row_idx}"].value = source_lookup[ord_id].get(src_col)
            updated_count += 1
        else:
            missing_orders.append({"Row in DATA": row_idx, "Order ID": ord_id})

    # 5. Check for Complete Mismatch
    if updated_count == 0 and len(missing_orders) > 0:
        raise NoMatchingOrdersError(
            "0 orders matched between the two files. "
            "Please verify that the correct daily files and customer tab were selected."
        )

    # 6. Save Updated Target Workbook to Memory Buffer
    output_buffer = io.BytesIO()
    wb_target.save(output_buffer)
    output_buffer.seek(0)

    audit_report = {
        "updated_count": updated_count,
        "missing_orders": missing_orders,
        "duplicate_orders": duplicate_source_orders,
        "warnings": advisory_warnings
    }

    return output_buffer, audit_report